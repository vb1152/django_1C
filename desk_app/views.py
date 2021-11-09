from django.shortcuts import render
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, QueryDict
from django.urls import reverse

# all my imports
import requests
import time
import datetime
from datetime import timedelta
from .models import User, ICparams, ICparamsForm, ICgroups, IC_scu, SimpleTable, Prices_type, PriceListTable, Producers, ProducersForJSgrid 
from .models import SalesForAbc, SalesForAbcForm, SafetyStock, DataForProgressbar, EventsDateTime, Prices, FoldersForFiles
from collections import defaultdict 
from django.contrib import messages
from django.db.models import Sum, F
from django.core.exceptions import ObjectDoesNotExist
import xml.etree.ElementTree as ET
from decimal import *
from .logic import *
import pandas as pd
from openpyxl import load_workbook
import os
import xlsxwriter
from django.core.mail import EmailMessage


# Create your views here.
#@login_required
def index(request):
    '''Index function. All control buttons are displayed on the main side.''' 
    return render(request, 'desk_app/index.html')

@login_required
def IC(request):
    '''Write or update in db parameters of connection to working 1C database '''

    if request.method == 'POST':
        # udate existing data about IC connection in db 
        if ICparams.objects.first():
            instace_param = ICparams.objects.first()
            form = ICparamsForm(request.POST, instance=instace_param)

            if form.is_valid():

                url = form.cleaned_data['url']
                port_IC = form.cleaned_data['port_IC']
                base_IC_name = form.cleaned_data['base_IC_name']
                odata_url = form.cleaned_data['odata_url']
                sessionIC_login = form.cleaned_data['sessionIC_login'].encode('UTF-8')
                sessionIC_password = form.cleaned_data['sessionIC_password']
                form.save()

                return HttpResponseRedirect(reverse("desk_app:IC"), )
            else: 
                return HttpResponse("Якась помилка. Спробуйте знову.")        
        # if not data about IC connection in db  
        else: 
            form = ICparamsForm(request.POST)

            if form.is_valid():
                instance = form.save(commit=False)
                url = form.cleaned_data['url']
                port_IC = form.cleaned_data['port_IC']
                base_IC_name = form.cleaned_data['base_IC_name']
                odata_url = form.cleaned_data['odata_url']
                sessionIC_login = form.cleaned_data['sessionIC_login'].encode('UTF-8')
                sessionIC_password = form.cleaned_data['sessionIC_password']
                instance.save()
                
                return HttpResponseRedirect(reverse("desk_app:IC"), )
            else:
                return HttpResponse("Якась помилка. Спробуйте знову.")


    form_IC = ICparamsForm()

    return render(request, 'desk_app/IC.html', {
        "message": "The user is authorized",
        "form_IC": form_IC.as_p,
    })

@login_required
def delete_data(request): 
    ''' Database clearing function. Deleting product groups and deleting all SCUs (including deleting
        all data associated with these records via Foreign key)''' 

    if request.method == "POST":

        if request.session.get('groups_count', False):
            del request.session['groups_count'] # delete session data when clear db 

        all_groups_count = ICgroups.objects.all().count()

        all_scu = IC_scu.objects.all().count()
        
        if all_groups_count > 0: 
            all_groups_delete = ICgroups.objects.all().delete()
            all_scu_delete = IC_scu.objects.all().delete()

            DataForProgressbar.objects.filter(name='groups').delete() #delete data about groups count in db 


            if not EventsDateTime.objects.filter(description__icontains="груп").exists():
                groups = EventsDateTime(description = 'Видалення груп', datetime_data = datetime.datetime.now())
                groups.save()

            if not EventsDateTime.objects.filter(description__icontains="номенклатур").exists():
                scu = EventsDateTime(description = 'Видалення номенклатури', datetime_data = datetime.datetime.now())
                scu.save()

            return JsonResponse({"comment": f'Deleted {all_groups_delete} groups, {all_scu_delete} SCU'},
                                status=200)
        else:
            return JsonResponse({"comment": f'Recorded to db {all_groups_count} groups and {all_scu} SCU'}, 
                                status=200)

def show_groups(request):
   
    main_group = ICgroups.objects.get(group_name = "Товары")

    all = ICgroups.get_all_children(main_group)
    
    new_data = []
    for group in all:
        new_data.append({'name': group['name']})
        if IC_scu.objects.filter(scu_parent_guid = group['object']).count() > 0:
            scu = IC_scu.objects.filter(scu_parent_guid = group['object']).values_list('scu_name', 'scu_barcode')
            
            if len(scu) == 1:
                new_data.append({'name': scu[0][0], "barcode": scu[0][1]})
            elif len(scu) > 1:
                for s in scu:
                    new_data.append({'name': s[0], "barcode": s[1]})

    table = PriceListTable(new_data)
    return render(request, "desk_app/groups.html", {"table": table})

@login_required
def load_all_data(request):
    '''
    Function to load folders to db. Into model ICgroups
    '''

    if request.method == "POST":
        # check if groups SCU are in db 
        all_groups_count = ICgroups.objects.all().count()

        # if no groups in db - load all groups. 
        if all_groups_count <= 0:
            groups_all = get_data_from_IS("/Catalog_Номенклатура?$format=json&$filter=IsFolder eq true&$select=Description,Ref_Key,Parent_Key,IsFolder,Code")

            data_progressbar = DataForProgressbar(name='groups', count=len(groups_all))
            data_progressbar.save()

            k = 0
            for i in groups_all:

                description = i['Description']
                code_9dig = i['Code']
                guid_group = i['Ref_Key']
                parent_guid_str = i['Parent_Key']
                is_folder_data = i['IsFolder']

                group = ICgroups(group_name=description, group_code=code_9dig, group_guid=guid_group, 
                                    parent_guid_str=parent_guid_str, is_folder=is_folder_data)
                group.save()

                k +=1
                if k > 10000:
                    break

            all_groups_count = ICgroups.objects.all().count()

            if folders_parents_create_no_req() : # call function to assign father group id to every group

                # save data about date and time loading folders 
                load_groups = EventsDateTime.objects.get(description__icontains="груп")
                load_groups.description = 'Завантаження груп'
                load_groups.datetime_data = datetime.datetime.now()
                load_groups.save()
                
                return JsonResponse({"comment": f'Готово!! В базі збережено {all_groups_count} груп.'}, 
                                        status=200)
            else: 
                return JsonResponse({"comment": 'Помилка. Якась проблема з присвоєнням батьківських зв"язків груп товарів.'}, 
                                        status=404)
        else:
        
            return JsonResponse({"comment": f'There are {all_groups_count} folders in db. No new download required.'}, 
                                        status=200)


                
def check_new_folders(request):
    '''Function to check new folders in 1c db. Run by button 'Update folders' on  http://127.0.0.1:8000/data page.'''

    if request.method == "GET":
        # get count of all groups from db 
        all_groups_count = ICgroups.objects.all().count()

        if all_groups_count <= 0: # if no groups in db - send info message to user
            return JsonResponse({"comment": 'Помилка. В базу поки не завантажені папки групи Товари. Спочатку виконайте завантаження.'}, 
                                        status=404)
        else: # if there are groups in db  
            # get groups from IC 
            groups_all = get_data_from_IS("/Catalog_Номенклатура?$format=json&$filter=IsFolder eq true&$select=Description,Ref_Key,Parent_Key,IsFolder,Code")
            
            if all_groups_count == len(groups_all): # check if count of groups in db and count groups in IC is equal
                
                # save data about operation in event list 
                group_event = EventsDateTime.objects.get(description__icontains="груп")     
                group_event.description = "Оновлення груп"
                group_event.datetime_data = datetime.datetime.now()
                group_event.save()
                
                
                return JsonResponse({"comment": 'All folders are saved in the database. No new folders found.'}, 
                                        status=404)
            elif len(groups_all) > all_groups_count: # if groups in IC more than groups in db, it meens some new groups are created
                
                list_Ref_Key_IC = list(map(lambda x : x['Ref_Key'], groups_all))

                list_group_Ref_Key_db = ICgroups.objects.all().values_list('group_guid', flat=True)
                
                # find ref codes for new groups. 
                res = list(set(list_Ref_Key_IC) - set(list_group_Ref_Key_db))

                name_data = []
                for x in res: # create list of new folders with all info 
                    name_data.append(list(filter(lambda name: name['Ref_Key'] == x, groups_all)))
                    
                check_save = 0
                for y in name_data: # save data about new folders to db 
                    group = ICgroups(group_name=y[0]['Description'], group_code=y[0]['Code'], group_guid=y[0]['Ref_Key'], 
                                    parent_guid_str=y[0]['Parent_Key'], is_folder=y[0]['IsFolder'])
                    group.save()
                    check_save += 1

                check = 0 
                for z in name_data: # save parent_guid ForeignKey in db 
                    try:
                        parent = ICgroups.objects.get(group_guid = z[0]['Parent_Key'])  # find parend group in db 
                        child = ICgroups.objects.get(group_guid = z[0]['Ref_Key']) # find child group in db 
                        child.parent_guid = parent
                        child.save()
                        check += 1
                    except ObjectDoesNotExist:
                        continue
               
                if not EventsDateTime.objects.filter(description__icontains="груп").exists():
                    group_event = EventsDateTime(description = 'Завантаження груп', datetime_data = datetime.datetime.now())
                    group_event.save()
                else:
                    group_event = EventsDateTime.objects.get(description__icontains="груп")
                    group_event.description = "Оновлення груп"
                    group_event.datetime_data = datetime.datetime.now()
                    group_event.save()


                if len(name_data) == check == check_save: # if groups, saved with foreignkey connections, same amount in db 
                    return JsonResponse({"comment": f'Saved {check} new folders in to db'}, status=200)
                elif len(name_data) == check_save and check_save > check:
                    return JsonResponse({"comment": f'Error. Saved {check_save}, but tied {check} new folders'}, status=200)
                elif len(name_data) != check_save:
                    return JsonResponse({"comment": f'Error. Saved only {check_save}, but in 1c db finded {len(name_data)} new folders.'}, status=200)

def load_price_from_xml(request):
    ''' Function to load prices from xml file. Match faster than download prices via odata since 
        result of odata request is more than 13000 rows. Run by button 'Load prices from XML' button on .../data page.
    '''
    
    if request.method == "GET":

        tree = ET.parse('C:\\1cbitrix\\1cbitrix\\offers.xml')
        root = tree.getroot()

        for price in root.iter():
            if price.tag == "Предложение":
                try: 
                    scu_current = IC_scu.objects.get(scu_guid = price[0].text)

                    price = Prices(scu=scu_current, ret_price=Decimal(price[5][0][2].text), who_price=Decimal(price[5][1][2].text))
                    price.save()
                               
                except (ObjectDoesNotExist, IndexError):
                    continue    

        all_prices = Prices.objects.all().count()

        # save to db info about operation (description, date-time)
        event(event = "Завантаження", object = "цін з XML", xml_datetime = root.attrib["ДатаФормирования"])

        return JsonResponse({"comment": f'Done!! {all_prices} rows of prices are stored in the database.'}, status=200)


def event(**kwargs):
    ''' Help function to save info data about operation in db (load, updating prices, scu ets..)
    Need to track last operations in db. '''
    
    try:
        kwargs['xml_datetime']
    except KeyError:
        kwargs['xml_datetime'] = ""

    if not EventsDateTime.objects.filter(description__icontains=kwargs['object']).exists():
        scu_event = EventsDateTime(description = kwargs['event']+' '+kwargs['object']+' файл '+kwargs['xml_datetime'], datetime_data = datetime.datetime.now())
        scu_event.save()
        return True
    else:
        scu_event = EventsDateTime.objects.get(description__icontains=kwargs['object'])
        scu_event.description = kwargs['event'] + ' ' +  kwargs['object']+' файл '+kwargs['xml_datetime']
        scu_event.datetime_data = datetime.datetime.now()
        scu_event.save()

        return True
  

def delete_prices_xml(request): 
    ''' Functiont to delete prices from db. '''

    if request.method == "GET":
        prices = Prices.objects.all()
        rows = prices.delete()

        event(event = "Видалення", object = "цін з XML")

        return JsonResponse({"comment": f' {rows} price data have been removed from the database.'}, status=200)

def folders_parents_create_no_req():
    ''' Function to assign father group id to every group after download all groups '''

    all = ICgroups.objects.all() #get all groups from db

    for group in all:
        try:
            parent = ICgroups.objects.get(group_guid = group.parent_guid_str)  
            group.parent_guid = parent
            group.save()
        except ObjectDoesNotExist:
            continue
           
    main_group = ICgroups.objects.get(group_guid = "ab2bf0a5-7426-11e1-b2e4-0013d43ff493")
    main_group.in_stock = True
    main_group.save()

    return True


def progress_api(request):
    ''' Function to help create progress bar when saving folders to db.
    have to change it functionality to use request.session.
    '''

    if request.method == "GET":
        # groups_count = DataForProgressbar.objects.values_list('count', flat=True).get(name='groups')

        if request.session.get('groups_count', False):
            groups_count = request.session['groups_count']

        else:
            groups_count = DataForProgressbar.objects.values_list('count', flat=True).get(name='groups')
            request.session['groups_count'] = groups_count 

        all_groups_count = ICgroups.objects.all().count()
        progress = int(all_groups_count/groups_count * 100)

        return JsonResponse({"progress": progress}, status=200)



# function creates new catalog Товари in db
@login_required
def load_data_from_IC(request):
    ''' function creates new catalog Товари in db
    '''

    # dev stage. Якщо товарні групи завантажені в базу, інше завантаження не відбувається. 
    all_groups_count = ICgroups.objects.all().count()
    if all_groups_count > 0: 
        return JsonResponse({"comment": f'There are currently {all_groups_count} groups in the database. Filling is not required.'},
                             status=200)
    else: 

        #connecting to IC 
        url = ICparams.objects.filter()[:1].get()
        session = requests.Session()
        session.auth = (url.sessionIC_login.encode('UTF-8'), url.sessionIC_password)
        response = session.get(url)

        # if connection to IC is successful 
        if response.status_code == 200:

            # working with db and loading data
            # get only folders from IC 
            groups_all = get_data_from_IS("/Catalog_Номенклатура?$format=json&$filter=IsFolder eq true&$select=Description,Ref_Key,Parent_Key,IsFolder,Code")

            # очистити всю базу даних перед записом. 
            ICgroups.objects.all().delete()
            
            # if goods_folder doesnt exist in db, create it
            guid_Tovaru = 'ab2bf0a5-7426-11e1-b2e4-0013d43ff493'
            try:
                goods_folder = ICgroups.objects.get(group_guid=guid_Tovaru)
            except ICgroups.DoesNotExist:
                # find in Catalog_Номенклатура data about Товары folder and save it to db  
                for i in groups_all:
                    if i['Ref_Key'] == guid_Tovaru and i['IsFolder'] == True:
                        description = i['Description']
                        code_9dig = i['Code']
                        guid_group = i['Ref_Key']
                        guid_parent = i['Parent_Key']
                        is_folder_data = i['IsFolder']
                        group = ICgroups(group_name=description, group_code=code_9dig, group_guid=guid_group, 
                                        is_folder=is_folder_data)
                        group.save()
            
            # get data about Товары folder from db 
            goods_folder = ICgroups.objects.get(group_guid=guid_Tovaru)
            # get children of Товары folder from IC 
            groups1 = get_data_from_IS("/Catalog_Номенклатура?$format=json&$filter=Parent_Key eq guid'" + goods_folder.group_guid + "' and IsFolder eq true")

            # save data about children groups of Товары folder in db 
            for i in groups1:
                if i['Parent_Key'] == guid_Tovaru and i['IsFolder'] == True:
                    description = i['Description']
                    code_9dig = i['Code']
                    guid_group = i['Ref_Key']
                    guid_parent = ICgroups.objects.get(group_guid=i['Parent_Key'])
                    is_folder_data = i['IsFolder']

                    group = ICgroups(group_name=description, 
                                        group_code=code_9dig, 
                                        group_guid=guid_group, 
                                        parent_guid=guid_parent,
                                        is_folder=is_folder_data)
                    group.save()

            # get data from db about groups of 1 order 
            groups1_from_db = list(ICgroups.objects.filter(parent_guid=goods_folder.id).values_list('group_guid', flat=True))
            
            groups2_list_guid = [] # create new list to save guids from groups 2 level to use it in next for loop 
            # for every group in all list of groups 
            # if group.parent guid is in 
            for groups2 in groups_all:
                if groups2['Parent_Key'] in groups1_from_db:
                    description = groups2['Description']
                    code_9dig = groups2['Code']
                    guid_group = groups2['Ref_Key']
                    guid_parent = ICgroups.objects.get(group_guid=groups2['Parent_Key'])
                    is_folder_data = groups2['IsFolder']

                    groups2_list_guid.append(guid_group)

                    group = ICgroups(group_name=description, 
                                        group_code=code_9dig, 
                                        group_guid=guid_group, 
                                        parent_guid=guid_parent,
                                        is_folder=is_folder_data)
                    group.save()
                    
            groups3_list_guid = []
            for groups3 in groups_all:
                if groups3['Parent_Key'] in groups2_list_guid:
                    description = groups3['Description']
                    code_9dig = groups3['Code']
                    guid_group = groups3['Ref_Key']
                    guid_parent = ICgroups.objects.get(group_guid=groups3['Parent_Key'])
                    is_folder_data = groups3['IsFolder']

                    groups3_list_guid.append(guid_group)

                    group = ICgroups(group_name=description, 
                                        group_code=code_9dig, 
                                        group_guid=guid_group, 
                                        parent_guid=guid_parent,
                                        is_folder=is_folder_data)
                    group.save()

            groups4_list_guid = []
            for groups4 in groups_all:
                if groups4['Parent_Key'] in groups3_list_guid:
                    description = groups4['Description']
                    code_9dig = groups4['Code']
                    guid_group = groups4['Ref_Key']
                    guid_parent = ICgroups.objects.get(group_guid=groups4['Parent_Key'])
                    is_folder_data = groups4['IsFolder']

                    groups4_list_guid.append(guid_group)

                    group = ICgroups(group_name=description, 
                                        group_code=code_9dig, 
                                        group_guid=guid_group, 
                                        parent_guid=guid_parent,
                                        is_folder=is_folder_data)
                    group.save()

            # find groups 5 
            groups5_list_guid = []
            for groups5 in groups_all:
                if groups5['Parent_Key'] in groups4_list_guid:
                    description = groups5['Description']
                    code_9dig = groups5['Code']
                    guid_group = groups5['Ref_Key']
                    guid_parent = ICgroups.objects.get(group_guid=groups5['Parent_Key'])
                    is_folder_data = groups5['IsFolder']

                    groups5_list_guid.append(guid_group)

                    group = ICgroups(group_name=description, 
                                        group_code=code_9dig, 
                                        group_guid=guid_group, 
                                        parent_guid=guid_parent,
                                        is_folder=is_folder_data)
                    group.save()
    
            all_groups_count = ICgroups.objects.all().count()
                    
            return JsonResponse({"comment": f'Names, codes and group structure Товари updated! In db {all_groups_count} folders'},
                                status=200)
        else:
            return JsonResponse({"comment": 'Something went wrong :(. '}, status=404)

@login_required
def load_scu(request):
    '''Funtion to save SCU to db. Create Odata request to 1c db and save result to apps db to IC_scu model''' 
    
    if request.method == "GET":
        scu_count = IC_scu.objects.all().count()

        if scu_count > 0:
            return JsonResponse({"comment": f'In db {scu_count} SCU. No download required.'},
                                        status=200)
        else:
            #get all scu from IC 
            scu_all = get_data_from_IS("/Catalog_Номенклатура?$format=json&$filter=IsFolder eq false&$select=Ref_Key,Parent_Key,Code,Description,Артикул")

            # get all groups ref_keys from db, as list
            groups_refs = list(ICgroups.objects.all().values_list('group_guid', flat=True))
            
            for scu in scu_all:
                if scu['Parent_Key'] in groups_refs:
                    scu_name = scu['Description']
                    scu_code = scu['Code']
                    scu_guid = scu['Ref_Key']
                    scu_parent_guid = ICgroups.objects.get(group_guid=scu['Parent_Key'])
                    scu_article = scu['Артикул']

                    scu_item = IC_scu(scu_name=scu_name, scu_code=scu_code, scu_guid=scu_guid, 
                                        scu_parent_guid=scu_parent_guid, scu_article=scu_article)
                    scu_item.save()                    

            scu_count = IC_scu.objects.all().count()

            if not EventsDateTime.objects.filter(description__icontains="номенклатур").exists():
                scu_event = EventsDateTime(description = 'Завантаження номенклатури', datetime_data = datetime.datetime.now())
                scu_event.save()
            else:
                scu_event = EventsDateTime.objects.get(description__icontains="номенклатур")
                scu_event.description = "Завантаження номенклатури"
                scu_event.datetime_data = datetime.datetime.now()
                scu_event.save()

            return JsonResponse({"comment": f'{scu_count} SCU added to db'},
                                            status=200)

def simple_list(request):
    # TODO функція показує просто несортований список товарів з цінами, завантаженими з ODATA 

    queryset = IC_scu.objects.all()
    table = SimpleTable(queryset)
    return render(request, 'desk_app/ic_scu_list.html', {'table': table})

def price_list_all(request):

    # get Товары item id to use it in the future
    tovaru_id = ICgroups.objects.filter(group_name='Товары')
   
    # get all groups form db 
    groups_list =  list(ICgroups.objects.all().values('id', 'group_name', 'parent_guid_id'))#.exclude(group_name='Товары'))
    
    # get list of parent_id's from scu 
    scu_parent_id =  IC_scu.objects.all().values_list('scu_parent_guid',  flat=True).distinct()
    
    itemdict = defaultdict(list)
    for item in groups_list:
        itemdict[item['parent_guid_id']].append({item['id']: item['group_name']})
        
    # create list of lists
    data_to_grid = []
    # function to get list of scu by its parent code 
    def get_scu_list(parent_code):
        result = IC_scu.objects.filter(scu_parent_guid=parent_code).values_list('scu_parent_guid', 'scu_name', 'scu_article', 'scu_barcode', 'scu_prices__who_price', 'scu_prices__ret_price', 'scu_guid_data__abc_result')
        
        return result 

    # recursion function to get all children for item from list of dicts
    def printitem(id_item, depth=0):        
        for child in itemdict[id_item]:
            
            child_id = [key for key in child.keys()][0]
            name = child[child_id]
                        
            # append data to list  
            data_to_grid.append([child_id, name])
            # if group has children scu 
            if child_id in scu_parent_id:
                # get list of child elements from db
                child_scu = get_scu_list(child_id)
                for scu in child_scu:
                    # append scu after group to the list
                    data_to_grid.append(list(scu))
                               
            printitem([key for key in child.keys()][0], depth+1)        
    printitem(tovaru_id[0].id)

    # create list with dicts for table html
    data_witk_keys = []
    for item in data_to_grid:
        data_witk_keys.append({#"id_item":item[0], 
                    "name": item[1],
                    "articul": "" if len(item) == 2 else item[2], 
                    "barcode": "" if len(item) == 2 else item[3],
                    "wh_price": "" if len(item) == 2 else item[4],
                    "rt_price": "" if len(item) == 2 else item[5],
                    "abc": "" if len(item) == 2 else item[6]
                     })

    # працююча відповідь для заповнення даних js-grid на сторінці Прайс
    table = PriceListTable(data_witk_keys)
    return render(request, "desk_app/price_list.html", {"table": table})

@login_required
def produsers(request):
    ''' Function to work with produsers data: load from 1C, render to user '''

    if request.method == 'POST':
        # to update data about produsers from IC
        if request.POST.get("update_all_prod_data"):
            guid_list = Producers.objects.all().values_list('producer_guid', flat=True)
            # scu_refs = list(IC_scu.objects.all().values_list('scu_guid', flat=True))

            produsers_all = get_data_from_IS("/Catalog_Контрагенты?$format=json&$filter=IsFolder eq false&$select=Ref_Key,Description")
            
            new_produsers = [] # list of new produsers names
            for produser in produsers_all:
                # if produser is new, save it to db 
                if produser['Ref_Key'] not in guid_list:
                    produser_name = produser['Description']
                    produser_guid = produser['Ref_Key']
                    produser_data = Producers(producer_guid=produser_guid, producer_name=produser_name)
                    produser_data.save()

                    new_produsers.append('produser_name') # add data to list of new produsers names
                else:
                    # if produser is not new, update its name in db  
                    Producers.objects.filter(producer_guid=produser['Ref_Key']).update(producer_name=produser['Description'])

            # return HttpResponseRedirect(reverse("desk_app:produsers"))  
            if len(new_produsers) > 0:
                messages.add_message(request, messages.INFO, f'{len(new_produsers)} suppliers have been added to the database.',  extra_tags='0')

            return render(request, "desk_app/produsers.html")
        
        if request.POST.get("clean_files"): #save to db address with clean files from produsers
            try: # створюємо нову адресу, якщо в базі немає запису.  
                clean_files_address = FoldersForFiles.objects.all()[:1].get()
                clean_files_address.clean_files = str(request.POST.get("data_clean"))
                clean_files_address.save()

                return JsonResponse({"comment": 'New address for clean manufacturer files is:  '
                                            f'{request.POST.get("data_clean")}'},
                                            status=200)    

            except ObjectDoesNotExist:
                new_addrr = FoldersForFiles(clean_files = request.POST.get("data_clean"))
                new_addrr.save()

                return JsonResponse({"comment": 'An entry in the database has been created. Move all clean order files from vendors to '
                                        f'{str(request.POST.get("data_clean"))}'},
                                            status=200)
                    
        if request.POST.get("ready_files"): # save to db address ready-to-send files 
            try: # створюємо нову адресу, якщо в базі немає запису.  
                ready_files_address = FoldersForFiles.objects.all()[:1].get()
                ready_files_address.ready_files = str(request.POST.get("data_ready"))
                ready_files_address.save()

                return JsonResponse({"comment": 'New address for clean manufacturer files is: '
                                            f'{request.POST.get("data_ready")}'},
                                            status=200)    

            except ObjectDoesNotExist:
                new_addr_ready = FoldersForFiles(ready_files = request.POST.get("data_ready"))
                new_addr_ready.save()

                return JsonResponse({"comment": 'An entry in the database has been created. Move all clean order files from vendors to '
                                        f'{str(request.POST.get("data_ready"))}'},
                                            status=200)
           
    # check if some produsers is in db. If not - load data from 1C 
    check = Producers.objects.all().count()

    addr = FoldersForFiles.objects.all()[:1].get()

    if check == 0: 
        messages.add_message(request, messages.INFO, 'There are no suppliers in the database. Download?',  extra_tags='0')
    else:
        messages.add_message(request, messages.INFO, f'There are {check} suppliers in the database.',  extra_tags='1')
    
    return render(request, "desk_app/produsers.html", {'clean_adr': addr.clean_files, 'ready_adr': addr.ready_files})
 
@login_required
def produsers_api(request, **kwargs):
   
    if request.method == 'GET':# and request.is_ajax():
        # https://docs.djangoproject.com/en/3.2/ref/models/conditional-expressions/ 
        
        table = Producers.objects.all() \
            .filter(producer_name__contains = request.GET.get('Name_1C')) \
            .filter(short_prod_name__contains = request.GET.get('Short_name')) \
            .filter(contact_data__contains = request.GET.get('Contact')) \
            .filter(info_field__contains = request.GET.get('Info')) \
            .order_by('-active') \
            .values('id', 'producer_name', 'short_prod_name', 'contact_data', 'email_order', 'info_field', 'active', 'file_name', 'barcode_col', 'order_col', 'min_sum')
        
        return JsonResponse(list(table), safe=False)

    if request.method == 'PUT':
        new_data = QueryDict(request.body) # https://stackoverflow.com/questions/34024650/django-wsgirequest-object-has-no-attribute-put

        produser = Producers.objects.get(id=kwargs['id'])
        produser.short_prod_name = new_data['Short_name']
        produser.contact_data = new_data["Contact"]
        produser.email_order = new_data["Email"]
        produser.info_field = new_data["Info"]
        produser.active = eval(new_data["Active"].title())
        produser.file_name = new_data["file_name"]
        produser.barcode_col = new_data["barcode_col"]
        produser.order_col = new_data['order_col']
        produser.min_sum = new_data['min_sum']
        
        produser.save()
        return HttpResponse(status = 200)

    if request.method == 'POST':
        new_produser = Producers(short_prod_name = request.POST.get('Short_name'), 
                                contact_data = request.POST.get('Contact'), 
                                email_order = request.POST.get('Email'), 
                                info_field = request.POST.get('Info'))
        new_produser.save()
        return HttpResponse(status = 200)

    if request.method == 'DELETE':
        del_produser = Producers.objects.get(id=kwargs['id'])
        del_produser.delete()

        return HttpResponse(status = 200)

def load_barcode(request):

    if request.method == "GET":
        #get all scu from IC 
        barcode_all = get_data_from_IS("/InformationRegister_Штрихкоды?$format=json&$select=Штрихкод,Владелец")

        # get all groups ref_keys from db, as list
        scu_refs = list(IC_scu.objects.all().values_list('scu_guid', flat=True))

        # get starting time
        start = time.time()

        counter = 0

        for barcode in barcode_all:
            if barcode['Владелец'] in scu_refs:
                scu = IC_scu.objects.get(scu_guid = barcode['Владелец'])
                if scu.scu_barcode != barcode['Штрихкод']:
                    
                    if barcode['Штрихкод'].isspace():
                        continue
                    else: 
                        scu.scu_barcode = barcode['Штрихкод']
                        scu.save()
                        counter += 1
                
        # get time taken to run the for loop code 
        elapsed_time_fl = (time.time() - start)


        if not EventsDateTime.objects.filter(description__icontains="штрихкод").exists():
            barc = EventsDateTime(description = 'Завантаження штрихкодів', datetime_data = datetime.datetime.now())
            barc.save()
        else:
            barc = EventsDateTime.objects.get(description__icontains="штрихкод")
            barc.description = 'Оновлення штрихкодів'
            barc.datetime_data = datetime.datetime.now()
            barc.save()

        # повертаємось на головну сторінку 
        return JsonResponse({"comment": f'Завантажено {counter} штрихкодів. За {int(elapsed_time_fl)} секунд. '},
                                            status=200)
@login_required
def delete_scu(request):
    '''SCU removal function from the database'''
    
    if request.method == "GET":
        scu_all = IC_scu.objects.all()
        deleted = scu_all.delete()

        if not EventsDateTime.objects.filter(description__icontains="номенклатур").exists():
            del_scu = EventsDateTime(description = 'Видалення номенклатури', datetime_data = datetime.datetime.now())
            del_scu.save()
        else:
            del_scu = EventsDateTime.objects.get(description__icontains="номенклатур")
            del_scu.description = 'Видалення номенклатури'
            del_scu.datetime_data = datetime.datetime.now()
            del_scu.save()

        return JsonResponse({"comment": f'Видалено {deleted} SCU.'},
                                            status=200)    

@login_required
def delete_barcode(request):
    ''' Function to delete barcodes from db '''
    if request.method == "GET":

        # get starting time
        start = time.time()

        all_scu = IC_scu.objects.all()

        for scu in all_scu:
            scu.scu_barcode = None
            scu.save()
        
        # get time taken to run the for loop code 
        elapsed_time_fl = (time.time() - start)

        # повертаємось на головну сторінку 
        return JsonResponse({"comment": f'Deleted all barcodes for {int(elapsed_time_fl)} seconds.'},
                                            status=200)


@login_required
def load_prices(request):
    '''Function to load prices from via Odata to db'''


    # Save types of prices to a db (Оптовая, закупочная, Розница... )
    if Prices_type.objects.all().count() == 0:
        prices_types = get_data_from_IS("/Catalog_ТипыЦенНоменклатуры?$format=json")
        
        for price_type in prices_types:
            price_name = price_type['Description']
            price_code = price_type['Ref_Key']
            price_type_data = Prices_type(price_name=price_name, price_code=price_code)
            price_type_data.save()                    
    
    
    #get all scu from IC 
    prices_all = get_data_from_IS("/InformationRegister_ЦеныНоменклатуры_RecordType/SliceLast()?$format=json")

    # get all groups ref_keys from db, as list
    groups_refs = list(IC_scu.objects.all().values_list('scu_guid', flat=True))
    
    # get types of prices from db
    price_type = Prices_type.objects.all().values()
    
    counter = 0

    # get starting time
    start = time.time()

    for price in prices_all:
        if price['Номенклатура_Key'] in groups_refs:
            
            if price_type.get(price_code=price['ТипЦен_Key'])['price_name'] == 'Оптовая':
                IC_scu.objects.filter(scu_guid = price['Номенклатура_Key']).update(who_price=price['Цена'])
                counter+=1

            elif price_type.get(price_code=price['ТипЦен_Key'])['price_name'] == 'Розница':
                IC_scu.objects.filter(scu_guid = price['Номенклатура_Key']).update(ret_price=price['Цена'])        
                counter+=1
            
    # get time taken to run the for loop code 
    elapsed_time_fl = (time.time() - start) 
    
    return JsonResponse({"comment": f'Оновлено {counter} цін за {elapsed_time_fl} сек.'},
                                        status=200)
 
def data(request):
    '''Function to load /data.html page'''
    return render(request, "desk_app/data.html")

def data_api(request):
    '''Function to send responses to ajax requests from data.html page'''

    if request.method == "GET":

        folders_quantity = ICgroups.objects.all().count()
        scu_quantity = IC_scu.objects.all().count()
        barcode = IC_scu.objects.filter(scu_barcode__isnull=False).count()
        prices_count = Prices.objects.all().count()

        prices = Prices.objects.all().values()

        event_group  = EventsDateTime.objects.filter(description__icontains="груп").values_list("description", "datetime_data")
        event_scu = EventsDateTime.objects.filter(description__icontains="номенклатури").values_list("description", "datetime_data")

        event_barcode = EventsDateTime.objects.filter(description__icontains="штрих").values_list("description", "datetime_data")
        event_price = EventsDateTime.objects.filter(description__icontains="цін").values_list("description", "datetime_data")
        

    return JsonResponse({"folders_quantity": folders_quantity, "scu_quantity": scu_quantity, "barcode": barcode, "prices_count": prices_count,
                        "event_group": str(event_group[0][0] + " " + event_group[0][1].strftime('%d.%m.%Y %H-%M-%S')),
                        "event_scu": str(event_scu[0][0] + " " + event_scu[0][1].strftime('%d.%m.%Y %H-%M-%S')),
                        "event_barcode": str(event_barcode[0][0] + " " + event_barcode[0][1].strftime('%d.%m.%Y %H-%M-%S')), 
                        "event_price": str(event_price[0][0] + " " + event_price[0][1].strftime('%d.%m.%Y %H-%M-%S'))
                        }, status=200)

@login_required
def IC_connection(request, IC_connection):
    '''Function to check connection to 1C db. Get fetch request from layout.js'''

    # Filter request based on IC_connection parameter 
    if IC_connection == 'check': 
        # get data from db about 
        if ICparams.objects.first():
            
            url = ICparams.objects.first()
            data = ICparams.objects.filter()[:1].get()
            session = requests.Session()
            session.auth = (data.sessionIC_login.encode('UTF-8'), data.sessionIC_password)
            response = session.get(url)
            print(url)
            if response.status_code == 200:
                return HttpResponse(status=200)
        else:
            return HttpResponse('Параметри підключення до 1С не знайдені.', status=404)
    
def login_view(request):
    if request.method == "POST":

        # Attempt to sign user in
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)

        # Check if authentication successful
        if user is not None:
            login(request, user)

            #delete all past order files from its folder FoldersForFiles.ready_files address
            file_address = FoldersForFiles.objects.all()[:1].get()
            for i in os.listdir(file_address.ready_files):
                os.remove(file_address.ready_files + '\\' + i)

            return HttpResponseRedirect(reverse("desk_app:index"))
        else:
            return render(request, "desk_app/login.html", {
                "message": "Невірний логін або пароль!"
            })
    else:
        return render(request, "desk_app/login.html")

def logout_view(request):
    logout(request)

    return HttpResponseRedirect(reverse("desk_app:index"))


def register(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST["email"]

        # Ensure password matches confirmation
        password = request.POST["password"]
        confirmation = request.POST["confirmation"]
        if password != confirmation:
            return render(request, "desk_app/register.html", {
                "message": "Паролі мають співпадати."
            })

        # Attempt to create new user
        try:
            user = User.objects.create_user(username, email, password)
            user.save()
        except IntegrityError:
            return render(request, "desk_app/register.html", {
                "message": "Цей логін вже зайнятий іншим користувачем."
            })
        login(request, user)
        return HttpResponseRedirect(reverse("desk_app:index"))
    else:
        return render(request, "desk_app/register.html")

@login_required
def abc(request):
    '''Function to load, calculate and save to db ABC data. 
    read more: https://en.wikipedia.org/wiki/ABC_analysis
    '''

    if request.method == 'GET':
        form_dates = SalesForAbcForm()

        # fill form with data from current ABC analysis
        try: 
            form_dates.fields['start_salesData'].initial = SalesForAbc.objects.all().values('start_salesData').first()['start_salesData']
            form_dates.fields['end_salesData'].initial = SalesForAbc.objects.all().values('end_salesData').first()['end_salesData']
        except:
            # print("Немає дат в базі даних. Потрібно ввести нові дати. ")
            return render(request, 'desk_app/abc.html', {"form_dates": form_dates, 
                                                        "message": "No dates in db. Need to enter new dates."
                                                        })

        return render(request, 'desk_app/abc.html', {"form_dates": form_dates})                                                

    if request.method == 'POST':
        print('post')
        form = SalesForAbcForm(request.POST)

        if form.is_valid():
            
            start_date = form.cleaned_data['start_salesData']
            end_date = form.cleaned_data['end_salesData']

            sales_all = get_data_from_IS(f"/AccumulationRegister_Продажи/Turnovers(EndPeriod=datetime'{end_date}T23:59:59',StartPeriod=datetime'{start_date}T00:00:00',Dimensions='Номенклатура')?&$format=json")

            # очистити всю базу даних перед записом. 
            SalesForAbc.objects.all().delete()
            
            for sale in sales_all:
                try:
                    scu = IC_scu.objects.get(scu_guid=sale['Номенклатура_Key'])
                except ObjectDoesNotExist:
                    continue
                sale_row = SalesForAbc(scu=scu, quantity=sale['КоличествоTurnover'],
                                        cost=sale['СтоимостьTurnover'],
                                        start_salesData=start_date, 
                                        end_salesData=end_date)
                sale_row.save()
                            
            # count and save to db ABC classes of every scu 
            data = SalesForAbc.objects.aggregate(Sum('cost'))
            count_ABC = SalesForAbc.objects.all().values('id', 'cost').order_by('-cost')
            a = 0
            for item in count_ABC:
                pers = item['cost']/data['cost__sum'] * 100
                a = a + pers
                if a < 70: 
                    SalesForAbc.objects.filter(id=item['id']).update(abc_result="A")
                elif a >= 70 and a < 90:
                    SalesForAbc.objects.filter(id=item['id']).update(abc_result="B")
                else:
                    SalesForAbc.objects.filter(id=item['id']).update(abc_result="C")

            form_dates = SalesForAbcForm()
            form_dates.fields['start_salesData'].initial = SalesForAbc.objects.all().values('start_salesData').first()['start_salesData']
            form_dates.fields['end_salesData'].initial = SalesForAbc.objects.all().values('end_salesData').first()['end_salesData']

            return render(request, 'desk_app/abc.html', {"form_dates": form_dates})

    

@login_required
def abc_api(request):
    '''
    Function to response to ajax request from abc.html with ABC analysis
    '''
    if request.method == 'GET' and request.is_ajax():

        table = SalesForAbc.objects.all() \
                .filter(scu__scu_name__contains = request.GET.get('Name')) \
                .filter(abc_result__contains = request.GET.get('Class')) \
                .values(Name=F('scu__scu_name'), Sales=F('quantity'), Sum=F('cost'), Class = F('abc_result')).order_by('-cost')
       
        return JsonResponse(list(table), safe=False)

@login_required
def sheet(request):
    return render(request, 'desk_app/sheet.html')

def sheet_api(request):
    '''
    Function to create list of SCU with data about: SCU name, scu balance on the first date, receipt of goods, 
    sales, scu balance on the last date and scu class acordint to ABC calculation. SCU name and ABC data are from db, 
    else are from Odata request to 1C db. 
    
    As sales data is availible only from dev db, d_start d_finish are hardcoded with timedelta 7 days. 
    For deployment version need to add form with data. 
    '''
    if request.method == "GET":
        d_start = datetime.date.today() - timedelta(days = 373) # дата формування + 6 днів 
        d_finish = datetime.date.today() - timedelta(days = 367) # дати формування
        
        d_start_abc = d_start - timedelta(days = 120)
        d_finish_abc = d_finish - timedelta(days = 30)
        
        sales_all = get_data_from_IS(f"/AccumulationRegister_ТоварыВРознице/BalanceAndTurnovers(EndPeriod=datetime'{d_finish}T23:59:59',StartPeriod=datetime'{d_start}T00:00:00',"
                                        "Dimensions='Номенклатура')?&$select=КоличествоOpeningBalance,"
                                        "КоличествоReceipt,КоличествоExpense,КоличествоClosingBalance,Номенклатура_Key&$format=json")

        names = 0
        for item in sales_all:
            try: # get name of item from db 
                name = IC_scu.objects.get(scu_guid=item['Номенклатура_Key'])
                item['Назва'] = name.scu_name
            except ObjectDoesNotExist: 

                # if name of scu not in db - get name from IC database by odata отримати назву позиції з бази 1С 
                item['Назва'] = ic_data_scu(item['Номенклатура_Key'])
                names += 1 

            try: # get ABC class from db 
                class_ABC = SalesForAbc.objects.get(scu__scu_guid=item['Номенклатура_Key'])#.values('abc_result')#[0]['abc_result']
                item['class'] = class_ABC.abc_result
            except ObjectDoesNotExist: # if no data about class in db... 
                item['class'] = "ССС"
                
        return JsonResponse(sales_all, safe=False)

@login_required
def stock_api(request, **kwargs):
    '''Function to create dict for js grid on page http://127.0.0.1:8000/stock in respone to ajax request.
    And to add insurance reserves, multiplicity and max order data for SCU, to be able to create orders to produsers. 
    '''
    if request.method == 'DELETE':
        data = QueryDict(request.body)
        if SafetyStock.objects.filter(scu_id=data['id']).exists():
            
            ss_data = SafetyStock.objects.filter(scu_id=data['id'])
            ss_data.delete()

            return JsonResponse({"comment": 'Safety stock data delete successfully. '},
                                status=200)


    if request.method == 'PUT':
        new_data = QueryDict(request.body) # https://stackoverflow.com/questions/34024650/django-wsgirequest-object-has-no-attribute-put

        if new_data["articul"]: # перевірка, що редагуються дані товару, а не групи 
            scu = IC_scu.objects.get(id=kwargs['id'])

            # якщо страховий запас/або інша інфо на товар вже записаний в таблиці 
            if SafetyStock.objects.filter(scu_id=scu.id).exists(): 
                stock = SafetyStock.objects.get(scu_id=scu.id)
                stock.scu = scu
                stock.provider = Producers.objects.get(id=int(new_data['manuf']))
                stock.safe_stock = new_data['sfs']
                stock.multipl = new_data['mlt']
                stock.only_max = new_data['maxOrd']
                stock.stock_max = new_data['maxSTK']
                stock.save()
            else: 
                stock = SafetyStock()
                stock.scu = scu
                stock.provider = Producers.objects.get(id=int(new_data['manuf']))
                stock.safe_stock = Decimal(new_data['sfs'].strip() or 0) # https://stackoverflow.com/questions/2941681/how-to-make-int-parse-blank-strings/2941975
                stock.multipl = Decimal(new_data['mlt'].strip() or 0)
                stock.only_max = Decimal(new_data['maxOrd'].strip() or 0)
                stock.stock_max = Decimal(new_data['maxSTK'].strip() or 0)
                stock.save()
            
            return HttpResponse(status = 200)
    if request.method == "GET":
        # get Товары item id to use it in the future
        tovaru_id = ICgroups.objects.filter(group_name='Товары')

        # get all groups form db 
        groups_list =  list(ICgroups.objects.all().values('id', 'group_name', 'parent_guid_id'))#.exclude(group_name='Товары'))
        
        # get list of parent_id's from scu 
        scu_parent_id = IC_scu.objects.all().values_list('scu_parent_guid',  flat=True).distinct() # працюючий код 

        # сортуємо групи по ієрархічним зв"язкам 
        itemdict = defaultdict(list)
        for item in groups_list:
            itemdict[item['parent_guid_id']].append({item['id']: item['group_name']})

        # create list of lists
        data_to_grid = []
        # function to get list of scu by its parent code 
        def get_scu_list(parent_code):
            result = IC_scu.objects.filter(scu_parent_guid=parent_code).values( 'id', 
                                                                                'scu_parent_guid', 'scu_name', 
                                                                                'scu_article', 'scu_barcode', 
                                                                                'scu_guid_data__abc_result', 
                                                                                'scu_safety_stock__safe_stock',
                                                                                'scu_safety_stock__provider',
                                                                                'scu_safety_stock__multipl',
                                                                                'scu_safety_stock__only_max',
                                                                                'scu_safety_stock__stock_max'
                                                                                )
            return result     

        # recursion function to get all children for item from list of dicts
        def printitem(id_item, depth=0):        
            for child in itemdict[id_item]: # знаходимо в ієрархічному списку груп itemdict позицію з id_item і беремо її кожен child
                            
                child_id = [key for key in child.keys()][0] # отримуємо id з dict {3272: 'Группа Бакалея. Хлеб'} 
                name = child[child_id] # отримуємо ім"я з використовуючи id {3272: 'Группа Бакалея. Хлеб'}  

                # append data to list  
                data_to_grid.append({'id': child_id, 'scu_name': name}) # додаємо id та name в список 

                # if group has children scu 
                if child_id in scu_parent_id: # якщо id групи є в списку груп, які мають дочірні елементи scu  
                    # get list of child elements from db
                    child_scu = get_scu_list(child_id)
                    for scu in child_scu:
                        data_to_grid.append(scu) # append scu after group to the list
                                
                printitem([key for key in child.keys()][0], depth+1)        

        printitem(tovaru_id[0].id)    

    return JsonResponse(data_to_grid, safe=False)

def stock(request):
    return render(request, 'desk_app/stock.html')

# rout to get list of manufacturers from db for /stock table
def manuf_api(request):
    if request.method == "GET":
        result = list(Producers.objects.filter(active = True).values('id', 'producer_name'))

        # завантаження даних по поточних залишках з 1С 
        sheet_for_orders(request)

        return JsonResponse(result, safe=False)

def orders(request):
    if request.method == "GET":
        return render(request, 'desk_app/orders.html')

def sheet_for_orders(request):
    '''When user open http://127.0.0.1:8000/orders - get data about sales from 1C 
    Data about sales are 2 years old for development version.
    '''

    d_start = datetime.date.today() - timedelta(days = 673) # дата формування + 6 днів 
    d_finish = datetime.date.today() - timedelta(days = 667)

    # Code for deploy version -> 2021-09-18T23:59:59   2021-09-12T00:00:00
    #d_start = datetime.date.today() - timedelta(days = 7)
    #d_finish = datetime.date.today() - timedelta(days = 1)

    sales_all = get_data_from_IS(f"/AccumulationRegister_ТоварыВРознице/BalanceAndTurnovers(EndPeriod=datetime'{d_finish}T23:59:59',StartPeriod=datetime'{d_start}T00:00:00',"
                                        "Dimensions='Номенклатура')?&$select=КоличествоOpeningBalance,"
                                        "КоличествоReceipt,КоличествоExpense,КоличествоClosingBalance,Номенклатура_Key&$format=json")
    
    # save data of odata request в request.session  
    request.session['sales_all'] = sales_all
     
    # delete request.session data whet web brouser closed 
    request.session.set_expiry(0)
    
    pass


def make_order_api(request):
    '''Function to calculate orders and create OR fill existing file for produser. '''

    sales_and_stock = request.session['sales_all']

    # отримуємо список товарів, по яким вказано цього постачальника 
    scu_from_safetystock = SafetyStock.objects.filter(provider = request.GET.get("id")).values("scu__scu_name", "scu__scu_barcode", "safe_stock", "scu__scu_guid_data__abc_result", "multipl", "only_max", "scu__scu_prices__who_price", "scu__scu_guid")

    if len(scu_from_safetystock) <= 0:
        return JsonResponse({"comment": 'В списку страхових запасів немає товарів із страховим запасом більше 0, що постачається цим постачальником.'},
                                status=400)
    else:
        order_sum = 0

        for item in scu_from_safetystock:

            # шукаємо дані по продажам у відомості по роздрібним продажам 
            # у вигляді(залишок на початок, приход, продажі, залишок на кінець, scu_guid)
            stock = [element for element in sales_and_stock if element['Номенклатура_Key'] == item['scu__scu_guid']]
            # підрахунок замовлення
            if len(stock) == 0: # не знайдено scu в обороті та на залишках за останні 7 днів

                if item['multipl'] == 0: # if multiplisity eq 0
                    item["order"] = item['safe_stock']
                    item['begin'] = 0
                    item['get'] = 0
                    item['sold'] = 0
                    item['in_stock'] = 0
                    item['scu__scu_guid'] = Decimal(item["order"] * item['scu__scu_prices__who_price']).quantize(Decimal('.01'), rounding=ROUND_DOWN)

                else:
                    item["order"] = ((item['safe_stock'] // item['multipl']) + 1 ) * item['multipl']
                    item['begin'] = stock[0]['КоличествоOpeningBalance']
                    item['get'] = stock[0]['КоличествоReceipt']
                    item['sold'] = stock[0]['КоличествоExpense']
                    item['in_stock'] = stock[0]['КоличествоClosingBalance']

                    item['scu__scu_guid'] = Decimal(item["order"] * item['scu__scu_prices__who_price']).quantize(Decimal('.01'), rounding=ROUND_DOWN)
            
            else: # if scu finded in list from 1c with sales data 
                item["order"] = calculate_order(stock = stock[0], item = item)

                item['begin'] = stock[0]['КоличествоOpeningBalance']
                item['get'] = stock[0]['КоличествоReceipt']
                item['sold'] = stock[0]['КоличествоExpense']
                item['in_stock'] = stock[0]['КоличествоClosingBalance']

                item['scu__scu_guid'] = Decimal(item["order"] * item['scu__scu_prices__who_price']).quantize(Decimal('.01'), rounding=ROUND_DOWN)
            
        # create new list with only positive orders 
        order = [d for d in scu_from_safetystock if d.get("order") > 0]
        # calculate total sum of order 
        order_sum = sum(item['scu__scu_guid'] for item in order)

        headers = ['Name', 'Barcode', 'SafeStock', 'Class', 'Mult', 'Max', 'Price', 'Total', 'Order', 'Зал', 'Отр', 'sales', 'Зал']

        file = Producers.objects.filter(id = request.GET.get("id")).values('producer_name', 'file_name', 'barcode_col', 'order_col', 'min_sum', 'email_order')
        
        file_address = FoldersForFiles.objects.all()[:1].get()

        file_addr_clean = file_address.clean_files + '\\' + file[0]['file_name']

        request.session["email_order"] = file[0]['email_order']
        # check if all data needed for produsers file (file name, barcode column, order column) are in db
        if len(file[0]['file_name']) > 0 and len(file[0]['barcode_col']) > 0 and len(file[0]['order_col']) > 0:

            data_from_excel = pd.read_excel(file_addr_clean, usecols=file[0]['barcode_col'], names = 'A', header=None)
            # Start by opening the spreadsheet and selecting the main sheet
            workbook = load_workbook(filename=file_addr_clean)
            sheet = workbook.active

            for item in order:
                row = data_from_excel.loc[data_from_excel['A'] == int(item['scu__scu_barcode'])] # find in pandas dataframe from excel row where exect barcode is
                # Write what you want into a specific cell sheet["C1"] = "writing ;)"
                
                sheet[file[0]['order_col'] + str(row.index[0] + 1)] = item['order']
            
            # Save the spreadsheet
            end_addr = file_address.ready_files + '\\' + 'Замовлення_ ' + file[0]['file_name']
            workbook.save(filename = end_addr)
            
            request.session["order_file_created"] = end_addr # save file addr to open it via python in future

            return JsonResponse({"comment": f'Замовлення створено та збережено до файлу замовлення постачальника  {end_addr}', 
                                "path_to_file": f"{end_addr}", 
                                "order": order,
                                'headers': headers, 
                                'order_sum': order_sum,
                                'min_sum': file[0]['min_sum'], 
                                'prod_name': file[0]['producer_name'],
                                'email': file[0]['email_order']
                                },
                                status=200)
        else: # if no file for order in db - create new file. 
            prod_name = (file[0]['producer_name']).replace(" ", "")
            end_addr = file_address.ready_files + '\\' + 'Замовлення_' + prod_name + '.xlsx'
            # Create a workbook and add a worksheet.
            workbook = xlsxwriter.Workbook(file_address.ready_files + '\\' + f'Замовлення_{prod_name}.xlsx', {'strings_to_numbers': False})
            worksheet = workbook.add_worksheet()

            # Start from the first cell. Rows and columns are zero indexed.
            row = 0
            col = 0
            # Add a bold format to use to highlight cells.
            bold = workbook.add_format({'bold': True, 'num_format': '#,##0.00'})
            # Add a number format for cells with money.
            money = workbook.add_format({'num_format': '#,##0.00'})
            # add width to column with name 
            worksheet.set_column(0, 0, 35)
            worksheet.set_column(1, 1, 15)
            # Iterate over the data and write it out row by row.
            for item in order:
                worksheet.write(row, col,     item['scu__scu_name'])
                worksheet.write_string(row, col + 1, item['scu__scu_barcode'])
                worksheet.write_number(row, col + 2, item['order'])
                worksheet.write(row, col + 3, item['scu__scu_prices__who_price'], money)
                #worksheet.write(row, col + 4, item['scu__scu_prices__who_price'], money)
                worksheet.write_formula(row, 4, f'=D{row+1}*C{row+1}')

                row += 1

            # Write a total using a formula.
            worksheet.write(row, 0, 'Total')
            worksheet.write_formula(row, 4, f'=SUM(E1:E{len(order)})', bold)

            workbook.close()

            request.session["order_file_created"] = end_addr # save file addr to open it via python in future
            
            return JsonResponse({"comment": f'Замовлення створено та збережено до нового файлу {end_addr}', 
                                "path_to_file": end_addr, 
                                "order": order,
                                'headers': headers, 
                                'order_sum': order_sum,
                                'min_sum': file[0]['min_sum'], 
                                'prod_name': file[0]['producer_name'],
                                'email': file[0]['email_order']
                                },
                                status=200)
            
def calculate_order (*args, **kwargs):
    '''Help function to calculate every order'''

    stock = kwargs["stock"]
    item = kwargs["item"]

    # calculate order 
    order = Decimal(stock["КоличествоExpense"]) + item["safe_stock"] - Decimal(stock["КоличествоClosingBalance"])

    # if multiplicity - calculate accordingly 
    if item["multipl"] > 0:
        order = ((order // item["multipl"]) + 1 ) * item["multipl"]
    # if max order is in place and calculated order is more than max -> order = max_order
    if item["only_max"] > 0 and order > item["only_max"]:
        order = item["only_max"]

    return order

def open_order_file_api(request):
    '''Function to open file.xlsx'''
    if request.method == 'GET':
        os.startfile(request.session["order_file_created"])
        return HttpResponse(status = 200)

@login_required
def send_order_file_api(request):
    '''Function to send files with order information to produsers.'''

    if request.method == 'GET':
        # https://docs.djangoproject.com/en/3.2/topics/email/
        email = EmailMessage(
        'Order',
        'This is order file',
        'fopbilych@gmail.com',
        [request.session["email_order"], ],
        reply_to=[request.session["email_order"]],
        headers={'Message-ID': 'foo'},
        )

        email.attach_file(request.session["order_file_created"])

        if email.send():
            return JsonResponse({"comment": 'Order is successfully sended. '},
                                status=200)
        else:
            return JsonResponse({"comment": 'Fail! '},
                                status=404)